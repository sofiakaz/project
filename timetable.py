import openpyxl
import pprint
from itertools import product
import copy
from openpyxl.styles import PatternFill, Border, Side, Alignment


all_letters = 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЫЭЮЯабвгдеёжзийклмнопрстуфхцчшщыэюя'

pp = pprint.PrettyPrinter(indent=4)

xl_teachers = openpyxl.load_workbook("Расписание уроков 9.xlsx", read_only=True)
students_schedule_9 = xl_teachers.active

students_free_time = {9: {'Понедельник': None,
                          'Вторник': None,
                          'Среда': None,
                          'Четверг': None,
                          'Пятница': None},
                      10: {'Понедельник': None,
                          'Вторник': None,
                          'Среда': None,
                          'Четверг': None,
                          'Пятница': None}}

num_of_les = 0
days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница']
for row in students_schedule_9.iter_rows(min_row=2, min_col=2, values_only=True):
    num_les = [num for num in row]
    num_of_les += 1
    for i in range(0, 5):
        if not num_les[i]:
            if not students_free_time[9][days[i]]:
                students_free_time[9][days[i]] = []
            students_free_time[9][days[i]].append(num_of_les)

xl_teachers = openpyxl.load_workbook("Расписание уроков 10.xlsx", read_only=True)
students_schedule_10 = xl_teachers.active

num_of_les = 0
for row in students_schedule_10.iter_rows(min_row=2, min_col=2, values_only=True):
    num_les = [num for num in row]
    num_of_les += 1
    for i in range(0, 5):
        if not num_les[i]:
            if not students_free_time[10][days[i]]:
                students_free_time[10][days[i]] = []
            students_free_time[10][days[i]].append(num_of_les)

def create_basic_datasets(sheet):
    exams = []
    electives = {}
    list_of_teachers = {}

    subjects = [cell.value for cell in sheet[1][2:]]
    teachers = [cell.value for cell in sheet[2][2:]]
    col_index = 0

    for subject in subjects:
        if subject:
            list_of_teachers[subject] = teachers[col_index]
            exams.append(subject)
            col_index += 1

    for row in range(3, sheet.max_row + 1):
        student_number = str(sheet[row][0].value).split('.')[0]
        student_choices = []

        for i, cell in enumerate(sheet[row][2:], start=2):
            if cell.value == "+":
                student_choices.append("+")
            else:
                student_choices.append("None")

        for i, choise in enumerate(student_choices):
            if choise == "+":
                subject = exams[i]
                if subject not in electives:
                    electives[subject] = []
                electives[subject].append(student_number)

    return list_of_teachers, electives

def create_files(clas, best_variants):
    with open(f'{clas}_class.txt', 'w', encoding='utf-8') as results_file:
        ind = 0
        for index, variant in enumerate(best_variants, start=1):
            free_subjects = {subject: days for subject, days in variant.items()}
            combinations = list(product(*free_subjects.values()))
            unique_combinations = []
            for combination in combinations:
                if len(set(combination)) == len(combination):
                    unique_combinations.append(combination)
            for i, combination in enumerate(unique_combinations, start=1):
                ind += 1
                results_file.write(f'{ind} вариант:\n')
                for subject, day in zip(free_subjects.keys(), combination):
                    lessons = day.lstrip(all_letters)
                    day = day.strip(' ,12345678')
                    if day == "Вторник":
                        results_file.write(f'{subject} во {day.lower()}{lessons} урок\n')
                    elif day == "Понедельник" or day == "Четверг":
                        results_file.write(f'{subject} в {day.lower()}{lessons} урок\n')
                    else:
                        day = day[:-1]
                        results_file.write(f'{subject} в {day.lower()}у{lessons} урок\n')
                results_file.write('\n')

    print(f'Файл {clas}_class.txt создан!')

xl = openpyxl.open("ОГЭ.xlsx", read_only=True)
sheet = xl.active

list_of_teachers_9, electives_9 = create_basic_datasets(sheet)

xl = openpyxl.open("ЕГЭ.xlsx", read_only=True)
sheet = xl.active

list_of_teachers_10, electives_10 = create_basic_datasets(sheet)

couples_9 = []

for subject_1, students_1 in electives_9.items():
    for subject_2, students_2 in electives_9.items():
        if subject_1 != subject_2 and not any(student in students_1 for student in students_2):
            couple = frozenset([subject_1, subject_2])
            for subject_3, students_3 in electives_9.items():
                if subject_3 not in couple and not any(student in students_3 for student in students_1) and not any(student in students_3 for student in students_2):
                    new_couple_1 = frozenset(list(couple) + [subject_3])
                    for subject_4, students_4 in electives_9.items():
                        if subject_4 not in new_couple_1 and not any(student in students_4 for student in students_1) and not any(student in students_4 for student in students_3) and not any(student in students_4 for student in students_2):
                            new_couple_2 = frozenset(list(new_couple_1) + [subject_4])
                            if new_couple_2 not in couples_9:
                                couples_9.append(new_couple_2)
                            for subject_5, students_5 in electives_9.items():
                                if subject_5 not in new_couple_2 and not any(
                                        student in students_5 for student in students_1) and not any(
                                        student in students_5 for student in students_3) and not any(
                                        student in students_5 for student in students_2) and not any(
                                        student in students_5 for student in students_4):
                                    new_couple_3 = frozenset(list(new_couple_2) + [subject_5])
                                    if new_couple_3 not in couples_9:
                                        couples_9.append(new_couple_3)

couples_9 = [list(couple) for couple in couples_9]

xl_teachers = openpyxl.load_workbook("Нагрузка.xlsx", read_only=True)
teachers_schedule = xl_teachers.active
teachers_free_time = {'Понедельник': {},
                      'Вторник': {},
                      'Среда': {},
                      'Четверг': {},
                      'Пятница': {}
                      }
current_day = None
current_teacher = None
first_index = 1
last_index = 8
for row in teachers_schedule.iter_rows(min_row=3, values_only=True):
    for day in teachers_free_time.keys():
        teachers_free_time[day][row[0]]=[]
        el_i = 0
        for element in row[first_index:last_index+1]:
            el_i += 1
            if not element:
                teachers_free_time[day][row[0]].append(el_i)
        first_index += 8
        last_index += 8
    first_index=1
    last_index=8

combinations_9 = {}
list_of_combinations_9 = []
count = 0
for couple_1 in couples_9:
    for couple_2 in couples_9:
        if couple_1 != couple_2 and not any(subject in couple_1 for subject in couple_2) and (len(couple_1)+len(couple_2)) == 9:
            combination = []
            combination.append(couple_1)
            combination.append(couple_2)
            combination = sorted(combination)
            if combination not in list_of_combinations_9:
                list_of_combinations_9.append(combination)
                count += 1
                combinations_9[count] = [couple_1, couple_2]
                flat_list = [item for sublist in combination for item in sublist]
                alone_subjects = []
                for subject in electives_9.keys():
                    if subject not in flat_list:
                        alone_subjects.append(subject)
                for couple in couples_9:
                    if all(subject in alone_subjects for subject in couple):
                        values_to_remove = couple
                        new_list = [item for item in alone_subjects if item not in values_to_remove]
                        combinations_9[count].append(couple)
                        alone_subjects = new_list
                combinations_9[count].extend(alone_subjects)

#чтобы вариантов было не бесконечное множество, выберем те варианты, где можно затратить минимальное
#количество учебных часов
best_variants_9 = []
for index, list_of_subjects in combinations_9.copy().items():
    comb_dict = {}
    subject_combinations = {}
    for day, schedule in teachers_free_time.items():
        for couple in list_of_subjects:
            if isinstance(couple, list):
                count = 0
                subjects_in_couple = []
                for subject in couple:
                    teacher = list_of_teachers_9[subject]
                    if students_free_time[9][day]:
                        if  str(students_free_time[9][day]).strip('[]') in str(schedule[teacher]).strip('[]'):
                            count += 1
                            subjects_in_couple.append(subject)
                if count == len(couple):
                    subjects_in_couple.sort()
                    subjects_key = ", ".join(subjects_in_couple)
                    if subjects_key not in subject_combinations:
                        subject_combinations[subjects_key] = []
                    subject_combinations[subjects_key].append(day + " " + str(students_free_time[9][day]).strip('[]'))
                count = 0
            else:
                subject = couple
                teacher = list_of_teachers_9[subject]
                if students_free_time[9][day]:
                    if str(students_free_time[9][day]).strip('[]') in str(schedule[teacher]).strip('[]'):
                        if subject not in comb_dict:
                            comb_dict[subject] = []
                        comb_dict[subject].append(day + " " + str(students_free_time[9][day]).strip('[]'))
    for subjects_key, days in subject_combinations.items():
        comb_dict[subjects_key] = days
    sorted_comb_dict = dict(sorted(comb_dict.items(), key=lambda item: len(item[1])))
    if len(comb_dict) == 3:
        best_variants_9.append(comb_dict)

couples_10 = []

for subject_1, students_1 in electives_10.items():
    for subject_2, students_2 in electives_10.items():
        if subject_1 != subject_2 and not any(student in students_1 for student in students_2):
            couple = frozenset([subject_1, subject_2])
            if couple not in couples_10:
                couples_10.append(couple)
            for subject_3, students_3 in electives_10.items():
                if subject_3 not in couple and not any(student in students_3 for student in students_1) and not any(student in students_3 for student in students_2):
                    new_couple = frozenset(list(couple) + [subject_3])
                    if len(new_couple) <= 3 and new_couple not in couples_10:
                        couples_10.append(new_couple)


couples_10 = [list(couple) for couple in couples_10]

combinations_10 = {}
list_of_combinations_10 = []
count = 0
for couple_1 in couples_10:
    for couple_2 in couples_10:
        if couple_1 != couple_2 and not any(subject in couple_1 for subject in couple_2):
            combination = []
            combination.append(couple_1)
            combination.append(couple_2)
            combination = sorted(combination)
            if combination not in list_of_combinations_10:
                list_of_combinations_10.append(combination)
                count += 1
                combinations_10[count] = [couple_1, couple_2]
                flat_list = [item for sublist in combination for item in sublist]
                alone_subjects = []
                for subject in electives_10.keys():
                    if subject not in flat_list:
                        alone_subjects.append(subject)
                for couple in couples_10:
                    if all(subject in alone_subjects for subject in couple):
                        values_to_remove = couple
                        new_list = [item for item in alone_subjects if item not in values_to_remove]
                        combinations_10[count].append(couple)
                        alone_subjects = new_list
                combinations_10[count].extend(alone_subjects)

#чтобы вариантов было не бесконечное множество, выберем те варианты, где можно затратить минимальное
#количество учебных часов
best_variants_10 = []
for index, list_of_subjects in combinations_10.copy().items():
    comb_dict = {}
    subject_combinations = {}
    for day, schedule in teachers_free_time.items():
        for couple in list_of_subjects:
            if isinstance(couple, list):
                count = 0
                subjects_in_couple = []
                for subject in couple:
                    teacher = list_of_teachers_10[subject]
                    if str(students_free_time[10][day]).strip('[]') in str(schedule[teacher]).strip('[]'):
                        count += 1
                        subjects_in_couple.append(subject)
                if count == len(couple):
                    subjects_in_couple.sort()
                    subjects_key = ", ".join(subjects_in_couple)
                    if subjects_key not in subject_combinations:
                        subject_combinations[subjects_key] = []
                    subject_combinations[subjects_key].append(day + ' ' + str(students_free_time[10][day]).strip('[]'))
                count = 0
            else:
                subject = couple
                teacher = list_of_teachers_10[subject]
                if str(students_free_time[10][day]).strip('[]') in str(schedule[teacher]).strip('[]'):
                    if subject not in comb_dict:
                        comb_dict[subject] = []
                    comb_dict[subject].append(day + ' ' + str(students_free_time[10][day]).strip('[]'))
    for subjects_key, days in subject_combinations.items():
        comb_dict[subjects_key] = days

    sorted_comb_dict = dict(sorted(comb_dict.items(), key=lambda item: len(item[1])))
    if len(comb_dict) == 3:
        best_variants_10.append((comb_dict))


create_files(9, best_variants_9)
create_files(10, best_variants_10)

with open('9_class.txt', 'r', encoding='utf-8') as f:
    all_variants_9 = f.readlines()

with open('10_class.txt', 'r', encoding='utf-8') as f:
    all_variants_10 = f.readlines()

variants_9 = []
current_variant_9 = None

for row in all_variants_9:
    if row.endswith('вариант:\n', 2):
        if current_variant_9:
            variants_9.append(current_variant_9)
        current_variant_9 = [row.rstrip('\n')]
    else:
        if current_variant_9 is not None and current_variant_9 != ' ':
            if row.rstrip('\n') != '':
                current_variant_9.append(row.rstrip('\n'))
variants_9.append(current_variant_9)

variants_10 = []
current_variant_10 = None

for row in all_variants_10:
    if row.endswith('вариант:\n', 2):
        if current_variant_10:
            variants_10.append(current_variant_10)
        current_variant_10 = [row.rstrip('\n')]
    else:
        if current_variant_10 is not None and current_variant_10 != ' ':
            if row.rstrip('\n') != '':
                current_variant_10.append(row.rstrip('\n'))
variants_10.append(current_variant_10)

free_classes = {'Понедельник': 0,
                'Вторник': 6,
                'Среда': 6,
                'Четверг': 4,
                'Пятница': 2}
day = {'Понедельник': [],
       'Вторник': [],
       'Среда': [],
       'Четверг': [],
       'Пятница': []}

ready_variants = []
for variant_1 in variants_9:
    cl_9 = copy.deepcopy(day)
    cl_10 = copy.deepcopy(day)
    ind = 0
    for elect_1 in variant_1[1:]:
        lessons_un_1 = elect_1.split()[0:-4]
        lesson_un_1 = ''
        lessons_1 = []
        if len(lessons_un_1) > 1:
            for lesson_1 in lessons_un_1:
                if lesson_1.endswith(',') and len(lesson_un_1) == 0:
                    lessons_1.append(lesson_1.rstrip(','))
                elif lesson_1.endswith(',') and len(lesson_un_1) > 0:
                    lesson_un_1 += lesson_1.rstrip(',')
                    lessons_1.append(lesson_un_1)
                    lesson_un_1 = ''
                elif lessons_un_1[-1] == lesson_1 and len(lesson_un_1) == 0:
                    lessons_1.append(lesson_1)
                elif lessons_un_1[-1] == lesson_1 and len(lesson_un_1) > 0:
                    lesson_un_1 += lesson_1
                    lessons_1.append(lesson_un_1)
                    lesson_un_1 = ''
                else:
                    lesson_un_1 += lesson_1
                    lesson_un_1 += ' '
        elif len(lessons_un_1) == 1:
            lessons_1 = lessons_un_1
        weekday_1 = elect_1.split(' ')[-3]
        les_1 = elect_1.split(' ')[-2]
        if weekday_1 == 'среду':
            weekday_1 = 'Среда'
        elif weekday_1 == 'пятницу':
            weekday_1 = 'Пятница'
        else:
            weekday_1 = weekday_1.capitalize()
        cl_9[weekday_1]=lessons_1
        for subject_1 in lessons_1:
            if subject_1 in list_of_teachers_9:
                teacher_1 = list_of_teachers_9[subject_1]
                day[weekday_1].append(teacher_1 + " " + subject_1 + " 9 класс" + " " + les_1 + " урок")
    saved_day = copy.deepcopy(day)
    for variant_2 in variants_10:
        ind += 1
        day = copy.deepcopy(saved_day)
        possibility = True
        for elect_2 in variant_2[1:]:
            lessons_un_2 = elect_2.split()[0:-5]
            lesson_un_2 = ''
            lessons_2 = []
            if len(lessons_un_2) > 1:
                for lesson_2 in lessons_un_2:
                    if lesson_2.endswith(',') and len(lesson_un_2) == 0:
                        lessons_2.append(lesson_2.rstrip(','))
                    elif lesson_2.endswith(',') and len(lesson_un_2) > 0:
                        lesson_un_2 += lesson_2.rstrip(',')
                        lessons_2.append(lesson_un_2)
                        lesson_un_2 = ''
                    elif lessons_un_2[-1] == lesson_2 and len(lesson_un_2) == 0:
                        lessons_2.append(lesson_2)
                    elif lessons_un_2[-1] == lesson_2 and len(lesson_un_2) > 0:
                        lesson_un_2 += lesson_2
                        lessons_2.append(lesson_un_2)
                        lesson_un_2 = ''
                    else:
                        lesson_un_2 += lesson_2
                        lesson_un_2 += ' '
            elif len(lessons_un_2) == 1:
                lessons_2 = lessons_un_2
            weekday_2 = elect_2.split(' ')[-4]
            if weekday_2 == 'среду':
                weekday_2 = 'Среда'
            elif weekday_2 == 'пятницу':
                weekday_2 = 'Пятница'
            else:
                weekday_2 = weekday_2.capitalize()
            cl_10[weekday_2]=lessons_2
            les_2 = elect_2.split(' ')[-3] + ' ' + elect_2.split(' ')[-2]
            for subject_2 in lessons_2:
                if subject_2 in list_of_teachers_10:
                    teacher = list_of_teachers_10[subject_2]
                    if teacher in saved_day[weekday_2]:
                        possibility = False
                        break
                    else:
                        day[weekday_2].append(list_of_teachers_10[subject_2]+ " " + subject_2 + " 10 класс" + " " + les_2 + " урок")
        num = 0
        sch_t = 0
        if possibility and day['Понедельник'] == []:
            for teachers in day.values():
                if any(element.startswith('Другова С.Г.') for element in teachers) and any(element.startswith('Нарицына Л.В.') for element in teachers):
                    num += 1
            for day_, teachers in day.items():
                if len(teachers) <= free_classes[day_]:
                    sch_t += 1
            if num > 1 and sch_t == 5:
                ready_variants.append(day)

    day = {'Понедельник': [],
           'Вторник': [],
           'Среда': [],
           'Четверг': [],
           'Пятница': []}

with open('Итоговый вариант.txt', 'w', encoding='utf-8') as random_result:
    for variant in ready_variants:
        for day, teachers in variant.items():
            random_result.write(f'{day}: {", ".join(teachers)}')
            random_result.write("\n")
        random_result.write("\n")

print("Файл Итоговый вариант.txt создан!")

file_path = 'Итоговый вариант.xlsx'

workbook = openpyxl.Workbook()
sheet = workbook.active

day_and_letter = {9: {'Понедельник': "B",
                       'Вторник': "D",
                       'Среда': "F",
                       'Четверг': "H",
                       'Пятница': "J"},
                  10: {'Понедельник': "C",
                       'Вторник': "E",
                       'Среда': "G",
                       'Четверг': "I",
                       'Пятница': "K"}}

def color(cell_range):
    gray_fill = PatternFill(start_color='BCBCBC', end_color='BCBCBC', fill_type='solid')
    for row_ in cell_range:
        for cell in row_:
            cell.fill = gray_fill


row = 1
for variant in ready_variants:
    cell_range = sheet[f'B{row+2}:B{row+9}']
    color(cell_range)
    cell_range = sheet[f'D{row + 2}:D{row + 9}']
    color(cell_range)
    cell_range = sheet[f'F{row + 2}:F{row + 9}']
    color(cell_range)
    cell_range = sheet[f'H{row + 2}:H{row + 9}']
    color(cell_range)
    cell_range = sheet[f'J{row + 2}:J{row + 9}']
    color(cell_range)
    cell_range = sheet[f'A{row}:K{row+9}']
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for row_ in cell_range:
        for cell in row_:
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
    sheet.row_dimensions[row].height = 30
    sheet.row_dimensions[row+1].height = 30
    sheet.row_dimensions[row+2].height = 50
    sheet.row_dimensions[row + 3].height = 50
    sheet.row_dimensions[row + 4].height = 30
    sheet.row_dimensions[row + 5].height = 30
    sheet.row_dimensions[row + 6].height = 30
    sheet.row_dimensions[row + 7].height = 30
    sheet.row_dimensions[row + 8].height = 80
    sheet.row_dimensions[row + 9].height = 90
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['G'].width = 18
    sheet.column_dimensions['H'].width = 18
    sheet.column_dimensions['I'].width = 18
    sheet.column_dimensions['J'].width = 18
    sheet.column_dimensions['K'].width = 18
    sheet.merge_cells(f'B{row}:C{row}')
    sheet.merge_cells(f'D{row}:E{row}')
    sheet.merge_cells(f'F{row}:G{row}')
    sheet.merge_cells(f'H{row}:I{row}')
    sheet.merge_cells(f'J{row}:K{row}')
    sheet[f'B{row}'] = 'Понедельник'
    sheet[f'D{row}'] = 'Вторник'
    sheet[f'F{row}'] = 'Среда'
    sheet[f'H{row}'] = 'Четверг'
    sheet[f'J{row}'] = 'Пятница'
    sheet[f'A{row+2}'] = '1'
    sheet[f'A{row+3}'] = '2'
    sheet[f'A{row+4}'] = '3'
    sheet[f'A{row+5}'] = '4'
    sheet[f'A{row+6}'] = '5'
    sheet[f'A{row+7}'] = '6'
    sheet[f'A{row+8}'] = '7'
    sheet[f'A{row+9}'] = '8'
    row += 1
    sheet[f'B{row}'] = '9 класс'
    sheet[f'C{row}'] = '10 класс'
    sheet[f'D{row}'] = '9 класс'
    sheet[f'E{row}'] = '10 класс'
    sheet[f'F{row}'] = '9 класс'
    sheet[f'G{row}'] = '10 класс'
    sheet[f'H{row}'] = '9 класс'
    sheet[f'I{row}'] = '10 класс'
    sheet[f'J{row}'] = '9 класс'
    sheet[f'K{row}'] = '10 класс'
    for day, less in variant.items():
        subjects_9 = []
        subjects_10 = []
        for lesson in less:
            clas = ''
            for el in lesson.split():
                if el == "10" or el == "9":
                    clas = el
            r_part = lesson.split('.')[-1]
            subject = []
            for p in r_part.split():
                if p != '9' and p != '10':
                    subject.append(p)
                else:
                    break
            subject = ' '.join(subject)
            if clas == '9':
                subjects_9.append(subject)
            else:
                subjects_10.append(subject)
        if students_free_time[9][day] and subjects_9:
            for les_num in students_free_time[9][day]:
                if len(subjects_9) == 1:
                    subject = ''.join(subjects_9)
                    sheet[f'{day_and_letter[9][day]}{row + les_num}'] = f'{subject.lower()}'
                else:
                    subjects = '/ '.join(subjects_9)
                    sheet[f'{day_and_letter[9][day]}{row + les_num}'] = f'{subjects.lower()}'
        if students_free_time[10][day] and subjects_10:
            for les_num in students_free_time[10][day]:
                if len(subjects_10) == 1:
                    subject = ''.join(subjects_10)
                    sheet[f'{day_and_letter[10][day]}{row + les_num}'] = f'{subject.lower()}'
                else:
                    subjects = '/ '.join(subjects_10)
                    sheet[f'{day_and_letter[10][day]}{row + les_num}'] = f'{subjects.lower()}'
    row += 15

workbook.save(file_path)
print("Файл Итоговый вариант.xlsx создан!")











